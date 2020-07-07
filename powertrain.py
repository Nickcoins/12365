from pymysql import connect
import sys
import prettytable as pt
from datetime import datetime
import openpyxl
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows


class CQ(object):
    def __init__(self):
        """创建connect链接"""
        try:
            print("正在连接数据库....")
            self.conn = connect(host='wardxu19858585.asuscomm.com', port=33066, user='guest', password='123456', database='test',
                           charset='utf8')
            self.cursor = self.conn.cursor()
            print("连接成功！")
        except:
            input("网络链接异常，退出程序.....")
            sys.exit()

    def __del__(self):
        """关闭cursor对象"""
        self.cursor.close()
        self.conn.close()

    def pretty_table(self, table_name, table):
        """使用prettytable 显示数据"""
        tb = pt.PrettyTable()
        tb.field_names = table_name
        for list_ in table:
            tb.add_row(list_)
        tb.add_column('序号', list(range(1, len(table)+1)), align="r")
        print(tb)

    def show_date(self):
        """输入日期"""
        print("-----输入日期范围-----")
        while True:
            try:
                sd = input("输入开始日期(yyyy/mm/dd)")
                ed = input("输入截止日期(yyyy/mm/dd)")
                start_date = datetime.strptime(sd, "%Y/%m/%d")
                stop_date = datetime.strptime(ed, "%Y/%m/%d")
                break
            except:
                print("输入格式错误")
        return [start_date, stop_date]

    def excute_sql(self, sql, show_date):
        """获取sql数据"""
        self.cursor.execute(sql, show_date)
        cols = self.cursor.description
        col = []
        for i in cols:
            col.append(i[0])  #获取表格列名
        self.pretty_table(col, self.cursor.fetchall())


    def show_detail_power_car(self, sql, show_date, num):
        """动力总成车型详单"""
        select_date = []
        select_date.extend(show_date)
        num_id = num
        self.cursor.execute(sql, show_date)
        complain = self.cursor.fetchall()[num_id - 1][1]
        select_date.append(complain)
        sql_2 = """SELECT f.`车型`, id1.`name` AS `抱怨分类`, id2.`name` AS `抱怨细节`,GROUP_CONCAT(DISTINCT (LEFT (f.`车辆配置`,4))) AS 车型年款, COUNT(f.`抱怨编号`) AS 抱怨数量 FROM `12365` AS f INNER JOIN id1 ON f.`抱怨分类`=id1.id INNER JOIN id2 ON f.`抱怨细节`=id2.id WHERE (f.`抱怨日期`>=%s AND f.`抱怨日期`<=%s) AND id1.`name`=%s and f.车企=%s AND id2.`name`=%s GROUP BY f.`车型`,f.`车企` ORDER BY 抱怨数量 DESC LIMIT 20"""
        self.excute_sql(sql_2, select_date)
        self.show_third_info(sql, show_date, sql_2, select_date, "show_detail_car_complain")

    def show_powertrain_compare(self, show_date):
        """1.动力总成抱怨数量"""
        sql = """select f.车企合并,
       count(if(id1.name='发动机',f.抱怨编号,null)) as 发动机,
       count(if(id1.name='变速器',f.抱怨编号,null)) as 变速器,
       count(f.抱怨编号) as 总计
from
    (select
           *,
           case
               when t.车企='上汽斯柯达' then '上汽大众'
               when t.车企 regexp '上汽通用' then '上汽通用'
               when t.车企='东风英菲尼迪' then '东风日产'
               when t.车企='广汽讴歌' then '广汽本田'
               when t.车企='一汽-大众奥迪' then '一汽-大众'
                   else t.车企
                       end as 车企合并
    from `12365` as t ) as f
inner join id1 on id1.id=f.抱怨分类
where id1.name regexp '发动机|变速器' and f.车企合并 regexp '上汽大众|上汽通用|广汽丰田|东风日产|广汽本田|一汽丰田|上汽集团|一汽-大众|东风本田' and f.`抱怨日期`>= %s AND f.`抱怨日期`<= %s
group by f.车企合并
order by 总计 desc"""
        self.excute_sql(sql, show_date)
        return sql

    def show_svwcomplain_trend(self, show_date):
        """2.动力总成每月数据"""
        sql = """select date_format(t.抱怨日期,'%%Y%%m') as 日期,
       count(if(id1.name='发动机',t.抱怨编号,null)) as 发动机,
       count(if(id1.name='变速器',t.抱怨编号,null)) as 变速器,
       count(if(id1.name regexp '发动机|变速器',null,t.抱怨编号)) as 其他
from `12365` as t
inner join id1 on id1.id=t.抱怨分类
where t.车企 regexp '上汽大众|斯柯达' and t.`抱怨日期`>= %s AND t.`抱怨日期`<= %s
group by 日期"""
        self.excute_sql(sql, show_date)
        return sql

    def show_svwengine_complain(self, show_date):
        """3.ea888/ea211 趋势"""
        sql = """select
    date_format(t.抱怨日期,'%%Y%%m') as 日期,
       count(if(t.发动机='EA888',t.抱怨编号,null)) as EA888,
       count(if(t.发动机='EA211',t.抱怨编号,null)) as EA211
from
     (select * ,
             case  when 车辆配置 regexp '1\\.8|2\\.0|300|320|330|380' then 'EA888'
                 when 车辆配置 regexp 'PHEV' then 'PHEV'
                 else 'EA211' end as `发动机`
     from `12365`as f where 车企 regexp '上汽大众|斯柯达' )as t
inner join id1 on id1.id=t.抱怨分类 
where t.`抱怨日期`>= %s AND t.`抱怨日期`<= %s and id1.name='发动机'
group by 日期
"""
        self.excute_sql(sql, show_date)
        return sql

    def show_powertrain_top20(self, show_date):
        """4.动力总成top20问题"""
        sql = """select
f.车企,f.车型,id1.name as 抱怨分类,id2.name as 抱怨细节,count(f.抱怨编号) as 数量
from
     (select
           *,
           case
               when t.车企='上汽斯柯达' then '上汽大众'
               when t.车企 regexp '上汽通用' then '上汽通用'
               when t.车企='东风英菲尼迪' then '东风日产'
               when t.车企='广汽讴歌' then '广汽本田'
               when t.车企='一汽-大众奥迪' then '一汽-大众'
                   else t.车企
                       end as 车企合并
    from `12365` as t ) as f
inner join id1 on id1.id=f.抱怨分类
inner join id2 on id2.id=f.抱怨细节
where  f.车企合并 regexp '上汽大众|上汽通用|广汽丰田|东风日产|广汽本田|一汽丰田|上汽集团|一汽-大众|东风本田'
  and f.`抱怨日期`>= %s AND f.`抱怨日期`<= %s
  and id1.name regexp '发动机|变速器'
group by f.车企, f.车型, id1.name, id2.name
order by 数量 desc limit 20"""
        self.excute_sql(sql, show_date)
        return sql

    def show_ea888_top20(self, show_date):
        """ea888 top20 抱怨"""
        sql = """select id2.name as 抱怨细节,count(t.抱怨编号) as 数量
from
    (select * ,
                 case  when 车辆配置 regexp '1\\.8|2\\.0|300|320|330|380' then 'EA888'
                     when 车辆配置 regexp 'PHEV' then 'PHEV'
                     else 'EA211' end as `发动机`
         from `12365`as f where 车企 regexp '上汽大众|斯柯达' )as t
inner join id1 on id1.id=t.抱怨分类
inner join id2 on id2.id=t.抱怨细节
where id1.name='发动机'
    and t.`抱怨日期`>= %s AND t.`抱怨日期`<= %s
    and t.发动机= 'EA888'
group by id2.name
order by 数量 desc limit 20"""
        self.excute_sql(sql, show_date)
        return sql

    def show_ea211_top20(self, show_date):
        """ea211 top20 抱怨"""
        sql = """select id2.name as 抱怨细节,count(t.抱怨编号) as 数量
from
    (select * ,
                 case  when 车辆配置 regexp '1\\.8|2\\.0|300|320|330|380' then 'EA888'
                     when 车辆配置 regexp 'PHEV' then 'PHEV'
                     else 'EA211' end as `发动机`
         from `12365`as f where 车企 regexp '上汽大众|斯柯达' )as t
inner join id1 on id1.id=t.抱怨分类
inner join id2 on id2.id=t.抱怨细节
where id1.name='发动机'
    and t.`抱怨日期`>= %s AND t.`抱怨日期`<= %s
    and t.发动机= 'EA211'
group by id2.name
order by 数量 desc limit 20"""
        self.excute_sql(sql, show_date)
        return sql

    def Gearbox(self, show_date):

        conn = connect(host='wardxu19858585.asuscomm.com', port=33066, user='guest', password='123456', database='test',
                       charset='utf8')
        sql = """select t.抱怨编号, t.车型, t.车辆配置, t.抱怨内容, t.抱怨日期, t.url, id2.name
             from `12365` as t
             inner join id1 on id1.id=t.抱怨分类
             inner join id2 on id2.id=t.抱怨细节
             where 车企 regexp '上汽大众|斯柯达'
                 and id1.name='变速器'
                 and t.`抱怨日期`>= '%s' AND t.`抱怨日期`<= '%s'""" % (str(show_date[0]), str(show_date[1]))
        d = pd.read_sql(sql, con=conn)
        gearbox_list = pd.read_excel(r"/Users/wardxu/Documents/GitHub/12365/变速箱匹配.xlsx")

        d['车辆配置'] = d['车辆配置'].str.replace(' ', '')
        gearbox_list['车辆配置'] = gearbox_list['车辆配置'].str.replace(' ', '')
        # print(d)
        # print(gearbox_list)
        a = pd.merge(d, gearbox_list, on=['车型', '车辆配置'], how='left')
        a['抱怨日期'] = pd.to_datetime(a['抱怨日期'])
        a['month'] = a['抱怨日期'].dt.to_period('M')
        # a = a.set_index('抱怨日期')
        # # print(a['变速箱型号'].resample('M').count().to_period('M'))
        # a = a.to_period('M')
        pd.set_option('display.max_columns', None)
        # print(a)
        #
        c = pd.pivot_table(a[['month', '变速箱型号']], index=['month'], columns=['变速箱型号'], aggfunc=len, fill_value=0)
        c.index = c.index.to_series().astype(str)

        gear_list = a['变速箱型号'].drop_duplicates(keep='last')

        for i in gear_list:
            temp = a[a.变速箱型号 == i]
            # print(temp.groupby('name')['抱怨编号'].count().sort_values(ascending=False))

        return c

    def save_report(self, show_date):
        """保存excel文件"""

        name = input("请输入保存文件名称：")
        file = openpyxl.Workbook()
        sheet_list = ['动力总成总体比较', '动力总成top20', '动力总成按月趋势', '发动机按月趋势', 'EA888 top20 问题', 'EA211 top20 问题']
        table_list = [self.show_powertrain_compare(show_date), self.show_powertrain_top20(show_date), self.show_svwcomplain_trend(show_date), self.show_svwengine_complain(show_date), self.show_ea888_top20(show_date), self.show_ea211_top20(show_date)]
        for n in range(len(sheet_list)):
            sheet_name = file.create_sheet(sheet_list[n])
            sql = table_list[n]
            self.cursor.execute(sql, show_date)
            cols = self.cursor.description
            col = []
            for i in cols:
                col.append(i[0])  # 获取表格列名
            table = self.cursor.fetchall()
            for i in range(len(col)):
                sheet_name.cell(row=1, column=i+1, value=col[i])
            for i in range(len(table)):
                for j in range(len(table[i])):
                    sheet_name.cell(row=i + 2, column=j + 1, value=table[i][j])
        sheet_gear = file.create_sheet('变速箱按月趋势')
        gear1 = self.Gearbox(show_date)
        # gear_cols = list(gear1)
        # for i in range(len(gear_cols)):
        #     sheet_gear.cell(row=1, column=i+2, value=gear_cols[i])
        # gear_index = gear1.index
        # for i in range(len(gear_index)):
        #     sheet_gear.cell(row=i+2, column=1, value=gear_index[i])
        # for i in range(gear1.shape[0]):
        #     for j in range(gear1.shape[1]):
        #         sheet_gear.cell(row=i+2, column=j+2, value=gear1.iloc[i, j])
        for r in dataframe_to_rows(gear1, index=True, header=True):
            print(r)
            sheet_gear.append(r)
        file.save(name + ".xlsx")

        print("保存成功")
        input("输入任意键继续：")



    def print_menu(self):
        """显示目录"""
        print("---------12365数据查询----------by wardxu version:11")
        sql_num = """SELECT COUNT(f.`抱怨编号`) FROM `12365` AS f"""
        self.cursor.execute(sql_num)
        num = self.cursor.fetchall()
        sql_date = """SELECT MAX(f.`抱怨日期`) FROM `12365` AS f"""
        self.cursor.execute(sql_date)
        date = self.cursor.fetchall()
        print("数据库总量：", num[0][0], "数据库最新日期：", date[0][0])
        print("6.动力总成抱怨对比")
        print("8.SVW抱怨趋势")
        print("9.EA888/EA211发动机趋势")
        print("10.EA888/EA211抱怨TOP20")
        print("11.保存报告")
        print("12.退出")
        return input("请输入选择的序号:")

    def run(self):
        """目录主程序"""
        #  错误循环
        while True:
            num = self.print_menu()
            if num == "1":
                # 查询车企数据
                show_date = self.show_date()
                self.show_all_brand(show_date)
            elif num == "2":
                # 查询车辆数据
                show_date = self.show_date()
                self.show_all_car(show_date)
            elif num == "3":
                # 查询抱怨分类数据
                show_date = self.show_date()
                self.show_all_complain(show_date)
            elif num == "4":
                #上汽大众车型数据
                show_date = self.show_date()
                print("-----选择大众/斯柯达-----")
                print("1.上汽大众")
                print("2.上汽斯柯达")
                print("3.返回")
                num_svw = input("请输入选择的序号:")
                if num_svw == "1":
                    show_date.extend(["上汽大众"])
                    self.show_allsvw_car(show_date)
                elif num_svw == "2":
                    show_date.extend(["上汽斯柯达"])
                    self.show_allsvw_car(show_date)
                elif num_svw == "3":
                    continue
                else:
                    print("输入有误请重新输入.....")
            elif num == "5":
                #查询动力总成数据
                show_date = self.show_date()
                print("-----选择动力总成分类-----")
                print("1.发动机")
                print("2.变速箱")
                print("3.动力电池")
                print("4.返回")
                num_power = input("请输入选择的序号:")
                if num_power == "1":
                    show_date.extend(["发动机"])
                    self.show_all_power(show_date)
                elif num_power == "2":
                    show_date.extend(["变速器"])
                    self.show_all_power(show_date)
                elif num_power == "3":
                    show_date.extend(["动力电池故障"])
                    self.show_all_battery(show_date)
                elif num_power == "4":
                    continue
                else:
                    print("输入有误请重新输入.....")
            elif num == "6":
                # 查询动力总成对比
                show_date = self.show_date()
                self.show_powertrain_compare(show_date)
            elif num == "7":
                # 查询动力总成top20
                show_date = self.show_date()
                print('-----选择发动机/变速箱-----')
                print("1.发动机")
                print("2.变速箱")
                print("3.返回")
                num_power = input("请输入选择的序号:")
                if num_power == "1":
                    show_date.extend(["发动机"])
                    self.show_powertrain_top20(show_date)
                elif num_power == "2":
                    show_date.extend(["变速器"])
                    self.show_powertrain_top20(show_date)
            elif num == "8":
                # 查询SVW动力总成趋势
                show_date = self.show_date()
                self.show_svwcomplain_trend(show_date)
            elif num == "9":
                # 查询EA888/EA211趋势
                show_date = self.show_date()
                self.show_svwengine_complain(show_date)
            elif num == "10":
                # 查询EA888/EA211top20
                show_date = self.show_date()
                print('-----选择EA888/EA211-----')
                print("1.EA888")
                print("2.EA211")
                print("3.返回")
                num_power = input("请输入选择的序号:")
                if num_power == "1":
                    show_date.extend(["EA888"])
                    self.show_ea888_top20(show_date)
                elif num_power == "2":
                    show_date.extend(["EA211"])
                    self.show_ea888_top20(show_date)
            elif num == '11':
                show_date = self.show_date()
                self.save_report(show_date)
            elif num == "12":
                break
            else:
                print("输入有误请重新输入.....")

def main():
    # 1.创建12365对象
    cq = CQ()
    # 2.调用对象的run方法，让其运行
    cq.run()

if __name__=='__main__':
    main()
