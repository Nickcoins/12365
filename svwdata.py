from pymysql import connect
import sys
import prettytable as pt
from datetime import datetime
import openpyxl

class CQ(object):
    def __init__(self):
        """创建connect链接"""
        try:
            print("正在连接数据库....")
            self.conn = connect(host='wardxu19858585.asuscomm.com', port=33066, user='guest', password='123456', database='test',
                           charset='utf8')
            self.cursor = self.conn.cursor()
            print("连接成功！")
        except:
            input("网络链接异常，退出程序.....")
            sys.exit()

    def __del__(self):
        """关闭cursor对象"""
        self.cursor.close()
        self.conn.close()

    def pretty_table(self, table_name, table):
        """使用prettytable 显示数据"""
        tb = pt.PrettyTable()
        tb.field_names = table_name
        for list_ in table:
            tb.add_row(list_)
        tb.add_column('序号', list(range(1, len(table)+1)), align="r")
        print(tb)

    def show_date(self):
        """输入日期"""
        print("-----输入日期范围-----")
        while True:
            try:
                sd = input("输入开始日期(yyyy/mm/dd)")
                ed = input("输入截止日期(yyyy/mm/dd)")
                start_date = datetime.strptime(sd, "%Y/%m/%d")
                stop_date = datetime.strptime(ed, "%Y/%m/%d")
                break
            except:
                print("输入格式错误")
        return [start_date, stop_date]

    def excute_sql(self, sql, show_date):
        """获取sql数据"""
        self.cursor.execute(sql, show_date)
        cols = self.cursor.description
        col = []
        for i in cols:
            col.append(i[0])  #获取表格列名
        self.pretty_table(col, self.cursor.fetchall())



    def show_all_brand(self, show_date):
        """1.显示车企数据"""
        sql = """SELECT f.`车企`,COUNT(DISTINCT f.`抱怨编号`) AS "抱怨数量" FROM `12365` AS f WHERE f.`抱怨日期`>= %s AND f.`抱怨日期`<= %s GROUP BY f.`车企` ORDER BY 抱怨数量 DESC LIMIT 20"""
        self.excute_sql(sql, show_date)
        self.show_second_info(sql, show_date, "show_detail_brand")

    def show_all_car(self, show_date):
        """2.显示车型数据"""
        sql = """SELECT f.`车型`,COUNT(DISTINCT f.`抱怨编号`) AS "抱怨数量" FROM `12365` AS f WHERE f.`抱怨日期`>= %s AND f.`抱怨日期`<= %s GROUP BY f.`车型` ORDER BY 抱怨数量 DESC LIMIT 20"""
        self.excute_sql(sql, show_date)
        self.show_second_info(sql, show_date, "show_detail_car")

    def show_all_complain(self, show_date):
        """3.显示抱怨分类细节"""
        sql = """SELECT id1.`name` AS `抱怨分类`,id2.`name` AS `抱怨细节`,COUNT(f.`抱怨编号`) AS 抱怨数量 FROM `12365` AS f INNER JOIN id1 ON f.`抱怨分类`=id1.id INNER JOIN id2 ON f.`抱怨细节`=id2.id WHERE f.`抱怨日期`>= %s AND f.`抱怨日期`<= %s GROUP BY id1.`name`,id2.`name` ORDER BY 抱怨数量 DESC LIMIT 20"""
        self.excute_sql(sql, show_date)
        self.show_second_info(sql, show_date, "show_detail_complain")

    def show_allsvw_car(self, show_date):
        """4.显示上汽大众车型数据"""
        sql ="""SELECT f.`车型`,COUNT(DISTINCT f.`抱怨编号`) AS 抱怨数量 FROM `12365` AS f WHERE (f.`抱怨日期`>= %s AND f.`抱怨日期`<= %s) and f.`车企`= %s GROUP BY f.`车型` ORDER BY 抱怨数量 DESC"""
        self.excute_sql(sql, show_date)
        self.show_second_info(sql, show_date, "show_detail_car")

    def show_all_power(self, show_date):
        """5.显示动力总成细节"""
        sql ="""SELECT f.车企 AS `车企`,COUNT(f.`抱怨编号`) AS 抱怨数量 FROM `12365` AS f INNER JOIN id1 ON f.`抱怨分类`=id1.id INNER JOIN id2 ON f.`抱怨细节`=id2.id WHERE f.`抱怨日期`>= %s AND f.`抱怨日期`<= %s and id1.`name`=%s GROUP BY f.车企 ORDER BY 抱怨数量 DESC LIMIT 20"""
        self.excute_sql(sql, show_date)
        self.show_second_info(sql, show_date, "show_detail_power")

    def show_all_battery(self, show_date):
        """5.3显示动力电池故障"""
        sql ="""SELECT f.车企 AS `车企`,COUNT(f.`抱怨编号`) AS 抱怨数量 FROM `12365` AS f INNER JOIN id1 ON f.`抱怨分类`=id1.id INNER JOIN id2 ON f.`抱怨细节`=id2.id WHERE f.`抱怨日期`>= %s AND f.`抱怨日期`<= %s and id2.`name`=%s  GROUP BY f.车企 ORDER BY 抱怨数量 DESC LIMIT 20"""
        self.excute_sql(sql, show_date)
        self.show_second_info(sql, show_date, "show_detail_battery")

    def show_second_info(self, sql, show_date, detail):
        """下级菜单"""
        while True:
            try:
                if detail == "show_detail_car":
                    num_id = int(input("输入序号查询详细--车型--信息，输入任意字符返回上级："))
                    self.show_detail_car(sql, show_date, num_id)
                elif detail == "show_detail_brand":
                    num_id = int(input("输入序号查询详细--车企--信息，输入任意字符返回上级："))
                    self.show_detail_brand(sql, show_date, num_id)
                elif detail == "show_detail_complain":
                    num_id = int(input("输入序号查询详细--抱怨--信息，输入任意字符返回上级："))
                    self.show_detail_complain(sql, show_date, num_id)
                elif detail == "show_detail_power":
                    num_id = int(input("输入序号查询详细--动力总成--信息，输入任意字符返回上级："))
                    self.show_detail_power(sql, show_date, num_id)
                elif detail == "show_detail_battery":
                    num_id = int(input("输入序号查询详细--动力总成--信息，输入任意字符返回上级："))
                    self.show_detail_battery(sql, show_date, num_id)
            except:
                break

    def show_third_info(self, sql, show_date, sql_2, select_date, detail):
        """下下级菜单"""
        while True:
            try:
                if detail == "show_detail_car":
                    num_id = int(input("输入序号查询详细--车型--信息，输入任意字符返回上级："))
                    self.show_detail_car(sql_2, select_date, num_id)
                elif detail == "show_detail_car_complain":
                    num_id = int(input("输入序号查询详细--具体车辆抱怨--信息，输入任意字符返回上级："))
                    self.show_detail_car_complain(sql_2, select_date, num_id)
                elif detail == "show_detail_power_car":
                    num_id = int(input("输入序号查询详细--具体车型抱怨--信息，输入任意字符返回上级："))
                    self.show_detail_power_car(sql_2, select_date, num_id)
            except:
                self.excute_sql(sql, show_date) #返回上级菜单同时显示上级表格内容
                break

    def save_excel(self, sql, select_date, select):
        """保存excel文件"""
        while True:
            try:
                id = input("是否保存更详细数据(包含用户具体抱怨)（y/n）：")
                if id == "y":
                    name = input("请输入保存文件名称：")
                    sql_4 = """SELECT id1.`name` AS 抱怨分类,id2.`name` AS 抱怨细节,f.`车型`,f.`车辆配置`,f.`抱怨内容`,f.`抱怨日期`,f.`url`,detail.`text` FROM `12365` AS f INNER JOIN id1 ON f.`抱怨分类`=id1.id INNER JOIN id2 ON f.`抱怨细节`=id2.id left join detail on detail.`url`=f.`url` WHERE (f.`抱怨日期`>=%s AND f.`抱怨日期`<=%s) AND f.`车型`=%s AND (id1.`name`=%s AND id2.`name`=%s)"""
                    self.cursor.execute(sql_4, select)
                    table = self.cursor.fetchall()
                    file = openpyxl.Workbook()
                    sheet = file.active
                    for i in range(len(table)):
                        for j in range(len(table[i])):
                            sheet.cell(row=i + 1, column=j + 1, value=table[i][j])
                    file.save(name + ".xlsx")
                    print("保存成功")
                    input("输入任意键继续：")
                    self.excute_sql(sql, select_date)
                    break
                elif id == "n":
                    self.excute_sql(sql, select_date)
                    break
                else:
                    print("输入错误，请重试")
            except:
                break

    def show_detail_car(self, sql, show_date, num):
        """车型详细信息"""
        select_date = []
        select_date.extend(show_date)
        num_id = num
        self.cursor.execute(sql, show_date)
        car = self.cursor.fetchall()[num_id-1][0]
        select = select_date[0:2]   #选择车企时show_date列表中包含车企信息，取前两项时间内容
        select.append(car)
        sql_2 = """SELECT f.`车型`,id1.`name` AS `抱怨分类`,id2.`name` AS `抱怨细节`,GROUP_CONCAT(DISTINCT (LEFT (f.`车辆配置`,4))) AS 车型年款,COUNT(f.`抱怨编号`) AS 抱怨数量 FROM `12365` AS f INNER JOIN id1 ON f.`抱怨分类`=id1.id INNER JOIN id2 ON f.`抱怨细节`=id2.id WHERE (f.`抱怨日期`>= %s AND f.`抱怨日期`<= %s) and f.`车型`=%s GROUP BY id1.`name`,id2.`name` ORDER BY 抱怨数量 DESC LIMIT 20"""
        self.excute_sql(sql_2, select)
        self.show_third_info(sql, show_date, sql_2, select, "show_detail_car_complain")


    def show_detail_brand(self, sql, show_date, num):
        """车企详细信息"""
        select_date = []
        select_date.extend(show_date)
        num_id = num
        self.cursor.execute(sql, show_date)
        brand = self.cursor.fetchall()[num_id-1][0]
        select_date.append(brand)
        sql_2 = """SELECT f.`车型`,COUNT(DISTINCT f.`抱怨编号`) AS 不重复抱怨数 FROM `12365` AS f WHERE (f.`抱怨日期`>= %s AND f.`抱怨日期`<= %s) and f.`车企`= %s GROUP BY f.`车型` ORDER BY 不重复抱怨数 DESC"""
        self.excute_sql(sql_2, select_date)
        self.show_third_info(sql, show_date, sql_2, select_date, "show_detail_car")



    def show_detail_complain(self, sql, show_date, num):
        """抱怨详细信息"""
        select_date = []
        select_date.extend(show_date)
        num_id = num
        self.cursor.execute(sql, show_date)
        complain = self.cursor.fetchall()[num_id-1][0:2]
        select_date.extend(complain)
        sql_2 = """SELECT f.`车型`,f.`车企`, id1.`name` AS `抱怨分类`, id2.`name` AS `抱怨细节`, COUNT(f.`抱怨编号`) AS 抱怨数量 FROM `12365` AS f INNER JOIN id1 ON f.`抱怨分类`=id1.id INNER JOIN id2 ON f.`抱怨细节`=id2.id WHERE (f.`抱怨日期`>=%s AND f.`抱怨日期`<=%s) AND (id1.`name`=%s AND id2.`name`=%s) GROUP BY f.`车型`,f.`车企` ORDER BY 抱怨数量 DESC LIMIT 20"""
        self.excute_sql(sql_2, select_date)
        self.show_third_info(sql, show_date, sql_2, select_date, "show_detail_car")


    def show_detail_car_complain(self, sql, select_date, num):
        """最终详单"""
        select_date_2 = []
        select_date_2.extend(select_date)
        num_id = num
        self.cursor.execute(sql, select_date)
        car = self.cursor.fetchall()[num_id-1][0:3]
        select = select_date[0:2]   #选择车企时show_date列表中包含车企信息，取前两项时间内容
        select.extend(car)
        sql_3 = """SELECT id1.`name` AS 抱怨分类,id2.`name` AS 抱怨细节,f.`车型`,f.`车辆配置`,f.`抱怨内容`,f.`抱怨日期` FROM `12365` AS f INNER JOIN id1 ON f.`抱怨分类`=id1.id INNER JOIN id2 ON f.`抱怨细节`=id2.id WHERE (f.`抱怨日期`>=%s AND f.`抱怨日期`<=%s) AND f.`车型`=%s AND (id1.`name`=%s AND id2.`name`=%s)"""
        self.excute_sql(sql_3, select)
        self.save_excel(sql, select_date, select)

    def show_detail_power(self, sql, show_date, num):
        """动力总成抱怨清单"""
        select_date = []
        select_date.extend(show_date)
        num_id = num
        self.cursor.execute(sql, show_date)
        brand = self.cursor.fetchall()[num_id - 1][0]
        select_date.append(brand)
        sql_2 = """SELECT id1.`name` as `抱怨分类`, id2.`name` AS `抱怨内容`,COUNT(f.`抱怨编号`) AS 抱怨数量 FROM `12365` AS f INNER JOIN id1 ON f.`抱怨分类`=id1.id INNER JOIN id2 ON f.`抱怨细节`=id2.id WHERE f.`抱怨日期`>= %s AND f.`抱怨日期`<= %s and id1.`name`=%s and f.车企=%s GROUP BY id2.`name` ORDER BY 抱怨数量 DESC LIMIT 20"""
        self.excute_sql(sql_2, select_date)
        self.show_third_info(sql, show_date, sql_2, select_date, "show_detail_power_car")

    def show_detail_battery(self, sql, show_date, num):
        """电池抱怨清单"""
        select_date = []
        select_date.extend(show_date)
        num_id = num
        self.cursor.execute(sql, show_date)
        brand = self.cursor.fetchall()[num_id - 1][0]
        select_date.append(brand)
        sql_2 = """SELECT f.`车型`, id1.name as `抱怨分类`, id2.`name` AS `抱怨细节`, COUNT(f.`抱怨编号`) AS 抱怨数量 FROM `12365` AS f INNER JOIN id1 ON f.`抱怨分类`=id1.id INNER JOIN id2 ON f.`抱怨细节`=id2.id WHERE (f.`抱怨日期`>=%s AND f.`抱怨日期`<=%s) AND id2.`name`=%s and f.车企= %s  GROUP BY f.`车型`,f.`车企`,id1.name ORDER BY 抱怨数量 DESC LIMIT 20"""
        self.excute_sql(sql_2, select_date)
        self.show_third_info(sql, show_date, sql_2, select_date, "show_detail_car_complain")

    def show_detail_power_car(self, sql, show_date, num):
        """动力总成车型详单"""
        select_date = []
        select_date.extend(show_date)
        num_id = num
        self.cursor.execute(sql, show_date)
        complain = self.cursor.fetchall()[num_id - 1][1]
        select_date.append(complain)
        sql_2 = """SELECT f.`车型`, id1.`name` AS `抱怨分类`, id2.`name` AS `抱怨细节`,GROUP_CONCAT(DISTINCT (LEFT (f.`车辆配置`,4))) AS 车型年款, COUNT(f.`抱怨编号`) AS 抱怨数量 FROM `12365` AS f INNER JOIN id1 ON f.`抱怨分类`=id1.id INNER JOIN id2 ON f.`抱怨细节`=id2.id WHERE (f.`抱怨日期`>=%s AND f.`抱怨日期`<=%s) AND id1.`name`=%s and f.车企=%s AND id2.`name`=%s GROUP BY f.`车型`,f.`车企` ORDER BY 抱怨数量 DESC LIMIT 20"""
        self.excute_sql(sql_2, select_date)
        self.show_third_info(sql, show_date, sql_2, select_date, "show_detail_car_complain")

    def show_powertrain_compare(self, show_date):
        """1.动力总成抱怨数量"""
        sql = """select f.车企合并,
       count(if(id1.name='发动机',f.抱怨编号,null)) as 发动机,
       count(if(id1.name='变速器',f.抱怨编号,null)) as 变速器,
       count(f.抱怨编号) as 总计
from
    (select
           *,
           case
               when t.车企='上汽斯柯达' then '上汽大众'
               when t.车企 regexp '上汽通用' then '上汽通用'
               when t.车企='东风英菲尼迪' then '东风日产'
               when t.车企='广汽讴歌' then '广汽本田'
               when t.车企='一汽-大众奥迪' then '一汽-大众'
                   else t.车企
                       end as 车企合并
    from `12365` as t ) as f
inner join id1 on id1.id=f.抱怨分类
where id1.name regexp '发动机|变速器' and f.车企合并 regexp '上汽大众|上汽通用|广汽丰田|东风日产|广汽本田|一汽丰田|上汽集团|一汽-大众|东风本田' and f.`抱怨日期`>= %s AND f.`抱怨日期`<= %s
group by f.车企合并
order by 总计 desc"""
        self.excute_sql(sql, show_date)
        return sql

    def show_svwcomplain_trend(self, show_date):
        """2.动力总成每月数据"""
        sql = """select date_format(t.抱怨日期,'%%Y%%m') as 日期,
       count(if(id1.name='发动机',t.抱怨编号,null)) as 发动机,
       count(if(id1.name='变速器',t.抱怨编号,null)) as 变速器,
       count(if(id1.name regexp '发动机|变速器',null,t.抱怨编号)) as 其他
from `12365` as t
inner join id1 on id1.id=t.抱怨分类
where t.车企 regexp '上汽大众|斯柯达' and t.`抱怨日期`>= %s AND t.`抱怨日期`<= %s
group by 日期"""
        self.excute_sql(sql, show_date)
        return sql

    def show_svwengine_complain(self, show_date):
        """3.ea888/ea211 趋势"""
        sql = """select
    date_format(t.抱怨日期,'%%Y%%m') as 日期,
       count(if(t.发动机='EA888',t.抱怨编号,null)) as EA888,
       count(if(t.发动机='EA211',t.抱怨编号,null)) as EA211
from
     (select * ,
             case  when 车辆配置 regexp '1\\.8|2\\.0|300|320|330|380' then 'EA888'
                 when 车辆配置 regexp 'PHEV' then 'PHEV'
                 else 'EA211' end as `发动机`
     from `12365`as f where 车企 regexp '上汽大众|斯柯达' )as t
inner join id1 on id1.id=t.抱怨分类 
where t.`抱怨日期`>= %s AND t.`抱怨日期`<= %s and id1.name='发动机'
group by 日期
"""
        self.excute_sql(sql, show_date)
        return sql

    def show_powertrain_top20(self, show_date):
        """4.动力总成top20问题"""
        sql = """select
f.车企,f.车型,id1.name as 抱怨分类,id2.name as 抱怨细节,count(f.抱怨编号) as 数量
from
     (select
           *,
           case
               when t.车企='上汽斯柯达' then '上汽大众'
               when t.车企 regexp '上汽通用' then '上汽通用'
               when t.车企='东风英菲尼迪' then '东风日产'
               when t.车企='广汽讴歌' then '广汽本田'
               when t.车企='一汽-大众奥迪' then '一汽-大众'
                   else t.车企
                       end as 车企合并
    from `12365` as t ) as f
inner join id1 on id1.id=f.抱怨分类
inner join id2 on id2.id=f.抱怨细节
where  f.车企合并 regexp '上汽大众|上汽通用|广汽丰田|东风日产|广汽本田|一汽丰田|上汽集团|一汽-大众|东风本田'
  and f.`抱怨日期`>= %s AND f.`抱怨日期`<= %s
  and id1.name regexp '发动机|变速器'
group by f.车企, f.车型, id1.name, id2.name
order by 数量 desc limit 20"""
        self.excute_sql(sql, show_date)
        return sql

    def show_ea888_top20(self, show_date):
        """ea888 top20 抱怨"""
        sql = """select id2.name as 抱怨细节,count(t.抱怨编号) as 数量
from
    (select * ,
                 case  when 车辆配置 regexp '1\\.8|2\\.0|300|320|330|380' then 'EA888'
                     when 车辆配置 regexp 'PHEV' then 'PHEV'
                     else 'EA211' end as `发动机`
         from `12365`as f where 车企 regexp '上汽大众|斯柯达' )as t
inner join id1 on id1.id=t.抱怨分类
inner join id2 on id2.id=t.抱怨细节
where id1.name='发动机'
    and t.`抱怨日期`>= %s AND t.`抱怨日期`<= %s
    and t.发动机= 'EA888'
group by id2.name
order by 数量 desc limit 20"""
        self.excute_sql(sql, show_date)
        return sql

    def show_ea211_top20(self, show_date):
        """ea888 top20 抱怨"""
        sql = """select id2.name as 抱怨细节,count(t.抱怨编号) as 数量
from
    (select * ,
                 case  when 车辆配置 regexp '1\\.8|2\\.0|300|320|330|380' then 'EA888'
                     when 车辆配置 regexp 'PHEV' then 'PHEV'
                     else 'EA211' end as `发动机`
         from `12365`as f where 车企 regexp '上汽大众|斯柯达' )as t
inner join id1 on id1.id=t.抱怨分类
inner join id2 on id2.id=t.抱怨细节
where id1.name='发动机'
    and t.`抱怨日期`>= %s AND t.`抱怨日期`<= %s
    and t.发动机= 'EA211'
group by id2.name
order by 数量 desc limit 20"""
        self.excute_sql(sql, show_date)
        return sql

    def save_report(self, show_date):
        """保存excel文件"""

        name = input("请输入保存文件名称：")
        file = openpyxl.Workbook()
        sheet_list = ['动力总成总体比较', '动力总成top20', '动力总成按月趋势', '发动机按月趋势', 'EA888 top20 问题', 'EA211 top20 问题']
        table_list = [self.show_powertrain_compare(show_date), self.show_powertrain_top20(show_date), self.show_svwcomplain_trend(show_date), self.show_svwengine_complain(show_date), self.show_ea888_top20(show_date), self.show_ea211_top20(show_date)]
        for n in range(len(sheet_list)):
            sheet_name = file.create_sheet(sheet_list[n])
            sql = table_list[n]
            self.cursor.execute(sql, show_date)
            cols = self.cursor.description
            col = []
            for i in cols:
                col.append(i[0])  # 获取表格列名
            table = self.cursor.fetchall()
            for i in range(len(col)):
                sheet_name.cell(row=1, column=i+1, value=col[i])
            for i in range(len(table)):
                for j in range(len(table[i])):
                    sheet_name.cell(row=i + 2, column=j + 1, value=table[i][j])
        file.save(name + ".xlsx")
        print("保存成功")
        input("输入任意键继续：")



    def print_menu(self):
        """显示目录"""
        print("---------12365数据查询----------by wardxu version:11")
        sql_num = """SELECT COUNT(f.`抱怨编号`) FROM `12365` AS f"""
        self.cursor.execute(sql_num)
        num = self.cursor.fetchall()
        sql_date = """SELECT MAX(f.`抱怨日期`) FROM `12365` AS f"""
        self.cursor.execute(sql_date)
        date = self.cursor.fetchall()
        print("数据库总量：", num[0][0], "数据库最新日期：", date[0][0])
        print("1.车企数据TOP20")
        print("2.车型数据TOP20")
        print("3.抱怨数据TOP20")
        print("4.上汽大众车型数据")
        print("5.动力总成系统数据")
        print("6.动力总成抱怨对比")
        print("7.动力总成抱怨TOP20")
        print("8.SVW抱怨趋势")
        print("9.EA888/EA211发动机趋势")
        print("10.EA888/EA211抱怨TOP20")
        print("11.保存报告")
        print("12.退出")
        return input("请输入选择的序号:")

    def run(self):
        """目录主程序"""
        #  错误循环
        while True:
            num = self.print_menu()
            if num == "1":
                # 查询车企数据
                show_date = self.show_date()
                self.show_all_brand(show_date)
            elif num == "2":
                # 查询车辆数据
                show_date = self.show_date()
                self.show_all_car(show_date)
            elif num == "3":
                # 查询抱怨分类数据
                show_date = self.show_date()
                self.show_all_complain(show_date)
            elif num == "4":
                #上汽大众车型数据
                show_date = self.show_date()
                print("-----选择大众/斯柯达-----")
                print("1.上汽大众")
                print("2.上汽斯柯达")
                print("3.返回")
                num_svw = input("请输入选择的序号:")
                if num_svw == "1":
                    show_date.extend(["上汽大众"])
                    self.show_allsvw_car(show_date)
                elif num_svw == "2":
                    show_date.extend(["上汽斯柯达"])
                    self.show_allsvw_car(show_date)
                elif num_svw == "3":
                    continue
                else:
                    print("输入有误请重新输入.....")
            elif num == "5":
                #查询动力总成数据
                show_date = self.show_date()
                print("-----选择动力总成分类-----")
                print("1.发动机")
                print("2.变速箱")
                print("3.动力电池")
                print("4.返回")
                num_power = input("请输入选择的序号:")
                if num_power == "1":
                    show_date.extend(["发动机"])
                    self.show_all_power(show_date)
                elif num_power == "2":
                    show_date.extend(["变速器"])
                    self.show_all_power(show_date)
                elif num_power == "3":
                    show_date.extend(["动力电池故障"])
                    self.show_all_battery(show_date)
                elif num_power == "4":
                    continue
                else:
                    print("输入有误请重新输入.....")
            elif num == "6":
                # 查询动力总成对比
                show_date = self.show_date()
                self.show_powertrain_compare(show_date)
            elif num == "7":
                # 查询动力总成top20
                show_date = self.show_date()
                print('-----选择发动机/变速箱-----')
                print("1.发动机")
                print("2.变速箱")
                print("3.返回")
                num_power = input("请输入选择的序号:")
                if num_power == "1":
                    show_date.extend(["发动机"])
                    self.show_powertrain_top20(show_date)
                elif num_power == "2":
                    show_date.extend(["变速器"])
                    self.show_powertrain_top20(show_date)
            elif num == "8":
                # 查询SVW动力总成趋势
                show_date = self.show_date()
                self.show_svwcomplain_trend(show_date)
            elif num == "9":
                # 查询EA888/EA211趋势
                show_date = self.show_date()
                self.show_svwengine_complain(show_date)
            elif num == "10":
                # 查询EA888/EA211top20
                show_date = self.show_date()
                print('-----选择EA888/EA211-----')
                print("1.EA888")
                print("2.EA211")
                print("3.返回")
                num_power = input("请输入选择的序号:")
                if num_power == "1":
                    show_date.extend(["EA888"])
                    self.show_ea888_top20(show_date)
                elif num_power == "2":
                    show_date.extend(["EA211"])
                    self.show_ea888_top20(show_date)
            elif num == '11':
                show_date = self.show_date()
                self.save_report(show_date)
            elif num == "12":
                break
            else:
                print("输入有误请重新输入.....")

def main():
    # 1.创建12365对象
    cq = CQ()
    # 2.调用对象的run方法，让其运行
    cq.run()

if __name__=='__main__':
    main()
